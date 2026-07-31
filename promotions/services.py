import csv
import hashlib
import io
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from html import escape
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from django.conf import settings
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from .models import Promotion, PromotionImportRun, PromotionSource
from .slug_utils import build_ascii_slug


HEADER_SYNONYMS = {
    "row_key": [
        "id",
        "uid",
        "promo id",
        "promotion id",
        "код",
        "код акции",
        "артикул",
        "sku",
    ],
    "title": [
        "title",
        "name",
        "promotion",
        "promo",
        "название",
        "название акции",
        "акция",
        "предложение",
        "товар",
        "модель",
        "продукт",
    ],
    "badge": [
        "badge",
        "label",
        "tag",
        "метка",
        "лейбл",
        "фишка",
        "плашка",
    ],
    "promotion_type": [
        "type",
        "promo type",
        "promotion type",
        "тип акции",
        "тип",
        "механика акции",
    ],
    "summary": [
        "summary",
        "short description",
        "short text",
        "preview",
        "кратко",
        "краткое описание",
        "анонс",
        "описание для превью",
    ],
    "details": [
        "details",
        "description",
        "full description",
        "long description",
        "условия",
        "подробности",
        "подробное описание",
        "описание",
        "механика",
    ],
    "brand": [
        "brand",
        "бренд",
        "vendor",
        "производитель",
    ],
    "category": [
        "category",
        "категория",
        "product category",
        "группа",
        "направление",
    ],
    "color": [
        "color",
        "colour",
        "цвет",
        "цвет товара",
        "расцветка",
    ],
    "promo_code": [
        "promo code",
        "code",
        "coupon",
        "промокод",
        "код купона",
    ],
    "promo_price": [
        "promo price",
        "price",
        "промоцена",
        "цена",
        "цена акции",
    ],
    "benefit_value": [
        "benefit",
        "discount",
        "gift",
        "скидка подарок",
        "скидка / подарок",
        "скидка/ подарок",
        "скидка",
        "подарок",
        "выгода",
        "размер скидки",
    ],
    "cta_label": [
        "cta label",
        "button text",
        "button",
        "текст кнопки",
        "кнопка",
    ],
    "cta_url": [
        "link",
        "url",
        "cta url",
        "button url",
        "ссылка",
        "ссылка на акцию",
        "url кнопки",
    ],
    "start_date": [
        "start",
        "start date",
        "date start",
        "дата начала",
        "начало",
        "с",
    ],
    "end_date": [
        "end",
        "end date",
        "date end",
        "дата окончания",
        "окончание",
        "финиш",
        "конец",
        "по",
    ],
    "sort_order": [
        "sort",
        "sort order",
        "order",
        "порядок",
        "сортировка",
    ],
    "is_featured": [
        "featured",
        "important",
        "highlight",
        "хит",
        "важная",
        "выделить",
    ],
    "is_published": [
        "published",
        "show",
        "visible",
        "active",
        "показывать",
        "опубликовано",
        "активна",
        "публикация",
    ],
    "customer_name": [
        "фио клиента",
        "клиент",
        "имя клиента",
        "покупатель",
    ],
    "phone": [
        "номер телефона",
        "телефон",
        "phone",
    ],
    "salesperson": [
        "ответственный продавец",
        "продавец",
        "менеджер",
    ],
    "acquisition_method": [
        "способ приобретения",
        "способ покупки",
        "способ оплаты",
    ],
    "store": [
        "то",
        "магазин",
        "салон",
        "точка",
    ],
    "status": [
        "статус",
        "состояние",
    ],
    "comment": [
        "комментарий",
        "примечание",
        "коммент",
    ],
    "product_status": [
        "статус по товару",
        "статус товара",
        "наличие",
    ],
    "order_date": [
        "дата заказа",
        "дата",
        "order date",
    ],
}


BRAND_HINTS = (
    "Samsung",
    "Apple",
    "Xiaomi",
    "Honor",
    "Poco",
    "Infinix",
    "Tecno",
    "Huawei",
    "Realme",
    "OnePlus",
    "Nothing",
    "Polaris",
    "Dreame",
    "Trouver",
    "Roborock",
    "Teclast",
    "Vivo",
    "Eufy",
    "Anker",
    "Soundcore",
    "QCY",
    "Elari",
    "Wifit",
    "Яндекс",
    "DeLonghi",
    "Omni",
)

BRAND_CANONICAL_NAMES = {
    brand.casefold(): brand
    for brand in BRAND_HINTS
}
SOURCE_ROW_KEY_MAX_LENGTH = 180


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    unpublished: int = 0
    duplicates: int = 0
    warnings: list[str] = field(default_factory=list)


class ImportValidationError(RuntimeError):
    """The source was readable, but its rows are unsafe to import."""


def normalize_header(value):
    text = (value or "").strip().casefold().replace("ё", "е")
    text = re.sub(r"[^\w]+", " ", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip()


def normalize_inline_text(value):
    return re.sub(r"\s+", " ", str(value or "").strip())


def canonicalize_brand(value):
    brand = normalize_inline_text(value)
    if not brand:
        return ""
    return BRAND_CANONICAL_NAMES.get(brand.casefold(), brand)


def append_product_variant(title, color):
    clean_title = normalize_inline_text(title)
    clean_color = normalize_inline_text(color)
    if not clean_color or normalize_header(clean_color) in normalize_header(clean_title):
        return clean_title
    return f"{clean_title} — {clean_color}"


def parse_bool(value):
    normalized = normalize_header(value)
    if not normalized:
        return None
    if normalized in {"1", "true", "yes", "y", "да", "активна", "показывать"}:
        return True
    if normalized in {"0", "false", "no", "n", "нет", "скрыть", "неактивна"}:
        return False
    return None


def parse_date_value(value):
    raw_value = str(value or "").strip()
    if not raw_value:
        return None

    raw_value = re.sub(r"(?<=\d),(?=\d)", ".", raw_value)

    for date_format in (
        "%d.%m.%Y",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%m/%d/%Y",
        "%d.%m.%y",
    ):
        try:
            return datetime.strptime(raw_value, date_format).date()
        except ValueError:
            continue
    return None


def is_open_ended_date_condition(value):
    normalized = normalize_header(value)
    if not normalized or normalized in {"-", "—"}:
        return True
    return any(
        marker in normalized
        for marker in (
            "бессроч",
            "без срока",
            "до окончания",
            "до конца запас",
            "до исчерпания",
            "пока",
            "по наличию",
        )
    )


def validate_mapped_promotion_dates(row_number, raw_row, mapped_data):
    normalized_row = {
        normalize_header(key): str(value or "").strip()
        for key, value in raw_row.items()
        if key
    }
    start_raw = extract_value(normalized_row, "start_date")
    end_raw = extract_value(normalized_row, "end_date")
    start_date = mapped_data.get("start_date")
    end_date = mapped_data.get("end_date")

    if start_raw and not start_date:
        raise ImportValidationError(
            f"Строка {row_number}: не удалось распознать дату начала «{start_raw}»."
        )
    if end_raw and not end_date and not is_open_ended_date_condition(end_raw):
        raise ImportValidationError(
            f"Строка {row_number}: не удалось распознать дату окончания «{end_raw}»."
        )
    if start_date and end_date and end_date < start_date:
        raise ImportValidationError(
            f"Строка {row_number}: дата окончания {end_date:%d.%m.%Y} "
            f"раньше даты начала {start_date:%d.%m.%Y}."
        )


def parse_excel_date_value(value):
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        if 20000 <= float(value) <= 60000:
            try:
                converted = from_excel(value)
            except (TypeError, ValueError):
                return None
            if isinstance(converted, datetime):
                return converted.date()
            if isinstance(converted, date):
                return converted
        return None

    return parse_date_value(str(value))


def parse_int_value(value, default=0):
    raw_value = (value or "").strip()
    if not raw_value:
        return default
    try:
        return int(raw_value)
    except ValueError:
        return default


def clean_inline_html(text):
    value = (text or "").strip()
    if not value:
        return ""
    return escape(value).replace("\n", "<br>")


def load_payload(url):
    request = Request(
        url,
        headers={"User-Agent": "CorpPortal/1.0 (+https://localhost)"},
    )

    try:
        with urlopen(request, timeout=30) as response:
            payload = response.read()
            charset = response.headers.get_content_charset("utf-8")
    except (HTTPError, URLError) as exc:
        raise RuntimeError(f"Не удалось загрузить таблицу: {exc}") from exc

    return payload, charset


def decode_payload(payload, charset):
    text = payload.decode(charset, errors="replace")
    if text.startswith("\ufeff"):
        text = text.lstrip("\ufeff")
    return text


def format_cell_text(value):
    if value in (None, ""):
        return ""

    parsed_date = parse_excel_date_value(value)
    if parsed_date:
        return parsed_date.strftime("%d.%m.%Y")

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()

    if isinstance(value, int):
        return str(value)

    return str(value).strip()


def extract_value(normalized_row, field_name):
    for alias in HEADER_SYNONYMS.get(field_name, []):
        candidate = normalized_row.get(normalize_header(alias), "")
        if candidate:
            return candidate.strip()
    return ""


def detect_promotion_kind(
    *,
    action_type="",
    badge="",
    title="",
    summary="",
    details="",
    promo_price="",
    benefit_value="",
):
    primary_text = " ".join(
        normalize_header(format_cell_text(value))
        for value in (action_type, badge)
        if value not in (None, "")
    )
    descriptive_text = " ".join(
        normalize_header(format_cell_text(value))
        for value in (title, summary, details, benefit_value)
        if value not in (None, "")
    )

    if "предзаказ" in primary_text:
        return Promotion.KIND_PREORDER
    if "подар" in primary_text:
        return Promotion.KIND_GIFT

    if "подар" in descriptive_text or "дарим" in descriptive_text:
        return Promotion.KIND_GIFT
    if "предзаказ" in descriptive_text:
        return Promotion.KIND_PREORDER

    promo_markers = ("скид", "промоцен", "промо")
    if any(marker in primary_text or marker in descriptive_text for marker in promo_markers):
        return Promotion.KIND_PROMO_PRICE

    normalized_benefit = normalize_inline_text(benefit_value)
    has_structured_benefit = bool(
        re.fullmatch(r"[\d\s.,]+%?", normalized_benefit)
    )
    if normalize_inline_text(promo_price) or has_structured_benefit:
        return Promotion.KIND_PROMO_PRICE

    return ""


def find_row_override(source, normalized_row):
    source_rules = getattr(settings, "PROMOTION_IMPORT_ROW_OVERRIDES", {})
    rules = source_rules.get(source.spreadsheet_id, ())

    for rule in rules:
        expected_values = rule.get("match", {})
        if all(
            normalize_header(extract_value(normalized_row, field_name))
            == normalize_header(expected_value)
            for field_name, expected_value in expected_values.items()
        ):
            row_override = dict(rule.get("set", {}))
            if rule.get("skip"):
                row_override["_skip"] = True
            return row_override

    return {}


def infer_brand(*values):
    merged = " ".join(format_cell_text(value) for value in values if value not in (None, ""))
    for brand in BRAND_HINTS:
        if brand.casefold() in merged.casefold():
            return canonicalize_brand(brand)
    return ""


def fetch_source_rows(source):
    payload, charset = load_payload(source.csv_url)
    text = decode_payload(payload, charset)

    rows = list(csv.reader(io.StringIO(text)))
    if len(rows) < source.header_row:
        raise RuntimeError("В таблице нет указанной строки с заголовками.")

    header = rows[source.header_row - 1]
    data_rows = rows[source.header_row :]

    prepared_rows = []
    for index, values in enumerate(data_rows, start=source.header_row + 1):
        if not any(cell.strip() for cell in values):
            continue

        padded = values + [""] * max(0, len(header) - len(values))
        row = {
            (header[position].strip() or f"Колонка {position + 1}"): padded[position].strip()
            for position in range(len(header))
        }
        prepared_rows.append((index, row))

    return prepared_rows


def bound_source_row_key(value, fallback=""):
    semantic_value = str(value or "").replace("+", " plus ")
    semantic_fallback = str(fallback or "").replace("+", " plus ")
    slug = build_ascii_slug(semantic_value, fallback=semantic_fallback)
    if len(slug) <= SOURCE_ROW_KEY_MAX_LENGTH:
        return slug

    digest = hashlib.sha256(slug.encode("utf-8")).hexdigest()[:12]
    prefix_length = SOURCE_ROW_KEY_MAX_LENGTH - len(digest) - 1
    return f"{slug[:prefix_length].rstrip('-')}-{digest}"


def build_row_key(source, row_number, raw_row, normalized_row, title, brand, promo_code, start_date, end_date):
    explicit_key = extract_value(normalized_row, "row_key")
    if explicit_key:
        return bound_source_row_key(explicit_key, fallback=explicit_key)

    base_parts = [
        title,
        brand,
        promo_code,
        str(start_date or ""),
        str(end_date or ""),
    ]
    base = "-".join(part for part in base_parts if part).strip()
    slug = bound_source_row_key(base)
    if slug:
        return slug

    fallback = next((value for value in raw_row.values() if value.strip()), "")
    return bound_source_row_key(
        fallback,
        fallback=f"{source.pk or 'source'}-row-{row_number}",
    )


def is_repeated_header_row(normalized_row):
    """Отсекает заголовки секций, повторённые внутри одной Google-таблицы."""
    matched_fields = 0
    for field_name in (
        "title",
        "brand",
        "promo_price",
        "benefit_value",
        "start_date",
        "end_date",
    ):
        value = normalize_header(extract_value(normalized_row, field_name))
        aliases = {normalize_header(alias) for alias in HEADER_SYNONYMS[field_name]}
        if value and value in aliases:
            matched_fields += 1
    return matched_fields >= 3


def map_row_to_promotion(source, row_number, raw_row):
    normalized_row = {
        normalize_header(key): value.strip()
        for key, value in raw_row.items()
        if key
    }

    if is_repeated_header_row(normalized_row):
        return None

    row_override = find_row_override(source, normalized_row)
    if row_override.get("_skip"):
        return None

    base_title = extract_value(normalized_row, "title")
    if not base_title:
        return None
    color = row_override.get("color") or extract_value(normalized_row, "color")
    title = append_product_variant(base_title, color)

    action_type = extract_value(normalized_row, "promotion_type")
    summary = extract_value(normalized_row, "summary")
    details = extract_value(normalized_row, "details")
    promo_price = extract_value(normalized_row, "promo_price")
    benefit_value = extract_value(normalized_row, "benefit_value")
    brand = canonicalize_brand(extract_value(normalized_row, "brand")) or infer_brand(
        title,
        summary,
        details,
        promo_price,
        benefit_value,
    )
    category = extract_value(normalized_row, "category")
    promo_code = extract_value(normalized_row, "promo_code")
    cta_url = extract_value(normalized_row, "cta_url")
    cta_label = extract_value(normalized_row, "cta_label")
    badge = extract_value(normalized_row, "badge")
    promotion_kind = detect_promotion_kind(
        action_type=action_type,
        badge=badge,
        title=title,
        summary=summary,
        details=details,
        promo_price=promo_price,
        benefit_value=benefit_value,
    )
    start_date_raw = extract_value(normalized_row, "start_date")
    end_date_raw = extract_value(normalized_row, "end_date")
    start_date = parse_date_value(start_date_raw)
    end_date = parse_date_value(end_date_raw)
    is_featured = parse_bool(extract_value(normalized_row, "is_featured")) or False
    explicit_published = parse_bool(extract_value(normalized_row, "is_published"))
    sort_order = parse_int_value(extract_value(normalized_row, "sort_order"), default=row_number * 10)

    row_key = build_row_key(
        source,
        row_number,
        raw_row,
        normalized_row,
        title,
        brand,
        promo_code,
        start_date,
        end_date,
    )

    clean_raw_data = {
        key.strip(): value.strip()
        for key, value in raw_row.items()
        if key and value and value.strip()
    }
    if row_override.get("color"):
        clean_raw_data["Цвет (правило импорта)"] = color

    if not badge and action_type:
        badge = action_type
    if not badge and promotion_kind == Promotion.KIND_GIFT:
        badge = "Подарок"
    if not badge and promotion_kind == Promotion.KIND_PREORDER:
        badge = "Предзаказ"

    display_action_type = action_type or (
        badge
        if promotion_kind in {Promotion.KIND_GIFT, Promotion.KIND_PREORDER}
        else ""
    )

    if not summary:
        summary_parts = []
        if display_action_type:
            summary_parts.append(display_action_type.title())
        if brand:
            summary_parts.append(brand)
        summary_parts.append(title)
        if promo_price:
            summary_parts.append(f"Промоцена: {promo_price}")
        if benefit_value:
            benefit_label = "Подарок" if promotion_kind == Promotion.KIND_GIFT else "Скидка"
            summary_parts.append(f"{benefit_label}: {benefit_value}")
        if details:
            summary_parts.append(details)
        summary = ". ".join(part for part in summary_parts if part)
        summary = summary[:220]

    if details:
        details = clean_inline_html(details)
    else:
        detail_parts = []
        if brand:
            detail_parts.append(
                f"<p><strong>Бренд:</strong> {clean_inline_html(brand)}</p>"
            )
        if display_action_type:
            detail_parts.append(
                f"<p><strong>Тип акции:</strong> {clean_inline_html(display_action_type)}</p>"
            )
        detail_parts.append(f"<p><strong>Товар:</strong> {clean_inline_html(title)}</p>")
        if promo_price:
            detail_parts.append(
                f"<p><strong>Промоцена:</strong> {clean_inline_html(promo_price)}</p>"
            )
        if benefit_value:
            benefit_label = "Подарок" if promotion_kind == Promotion.KIND_GIFT else "Скидка"
            detail_parts.append(
                f"<p><strong>{benefit_label}:</strong> {clean_inline_html(benefit_value)}</p>"
            )
        if end_date_raw and not end_date:
            detail_parts.append(
                f"<p><strong>Окончание:</strong> {clean_inline_html(end_date_raw)}</p>"
            )
        details = "".join(detail_parts)

    if not badge and promo_code:
        badge = f"Промокод {promo_code}"

    return {
        "source": source,
        "source_row_key": row_key,
        "title": title,
        "promotion_kind": promotion_kind,
        "badge": badge,
        "summary": summary,
        "details": details,
        "brand": brand,
        "category": category,
        "promo_price": promo_price,
        "benefit_value": benefit_value,
        "promo_code": promo_code,
        "cta_label": cta_label,
        "cta_url": cta_url,
        "start_date": start_date,
        "end_date": end_date,
        "sort_order": sort_order,
        "is_featured": is_featured,
        "is_published": (
            explicit_published if explicit_published is not None else source.auto_publish_imported
        ),
        "raw_data": clean_raw_data,
        "imported_at": timezone.now(),
    }


def build_worksheet_rows(worksheet):
    values_rows = list(worksheet.iter_rows(values_only=True))
    if not values_rows:
        return []

    header_index = 0
    for index, row in enumerate(values_rows[:5]):
        normalized_values = {
            normalize_header(format_cell_text(value))
            for value in row
            if format_cell_text(value)
        }
        if "модель" in normalized_values and (
            "дата заказа" in normalized_values or "фио клиента" in normalized_values
        ):
            header_index = index
            break

    header = [format_cell_text(value) for value in values_rows[header_index]]
    data_rows = values_rows[header_index + 1 :]

    prepared_rows = []
    for row_number, values in enumerate(data_rows, start=header_index + 2):
        if not any(format_cell_text(value) for value in values):
            continue

        row = {}
        for position, header_value in enumerate(header):
            if not header_value:
                continue
            cell_value = values[position] if position < len(values) else ""
            row[header_value] = format_cell_text(cell_value)

        if row:
            prepared_rows.append((row_number, row))

    return prepared_rows


def summarize_counter(counter):
    return ", ".join(f"{key}: {count}" for key, count in counter.most_common() if key)


def map_worksheet_to_preorder_promotion(source, worksheet, sort_order):
    prepared_rows = build_worksheet_rows(worksheet)
    preorder_entries = []
    model_names = []
    statuses = Counter()
    acquisition_methods = Counter()
    parsed_dates = []

    for row_number, raw_row in prepared_rows:
        normalized_row = {
            normalize_header(key): value.strip()
            for key, value in raw_row.items()
            if key
        }

        model_name = extract_value(normalized_row, "title")
        customer_name = extract_value(normalized_row, "customer_name")
        phone = extract_value(normalized_row, "phone")
        salesperson = extract_value(normalized_row, "salesperson")
        acquisition_method = extract_value(normalized_row, "acquisition_method")
        store = extract_value(normalized_row, "store")
        status = extract_value(normalized_row, "status")
        comment = extract_value(normalized_row, "comment")
        product_status = extract_value(normalized_row, "product_status")
        order_date_raw = extract_value(normalized_row, "order_date")
        order_date = parse_excel_date_value(order_date_raw) or parse_date_value(order_date_raw)

        if not any(
            [
                model_name,
                customer_name,
                phone,
                salesperson,
                acquisition_method,
                status,
                comment,
                product_status,
            ]
        ):
            continue

        if model_name:
            model_names.append(model_name)
        if status:
            statuses[status] += 1
        if acquisition_method:
            acquisition_methods[acquisition_method] += 1
        if order_date:
            parsed_dates.append(order_date)

        preorder_entries.append(
            {
                "row_number": row_number,
                "model": model_name,
                "order_date": order_date.strftime("%d.%m.%Y") if order_date else order_date_raw,
                "customer_name": customer_name,
                "phone": phone,
                "salesperson": salesperson,
                "acquisition_method": acquisition_method,
                "store": store,
                "status": status,
                "comment": comment,
                "product_status": product_status,
            }
        )

    if not preorder_entries:
        return None

    unique_models = []
    seen_models = set()
    for model_name in model_names:
        normalized = normalize_header(model_name)
        if normalized and normalized not in seen_models:
            unique_models.append(model_name)
            seen_models.add(normalized)

    title = worksheet.title.strip()
    brand = infer_brand(title, *unique_models)
    status_summary = summarize_counter(statuses)
    acquisition_summary = summarize_counter(acquisition_methods)
    preview_models = ", ".join(unique_models[:3])

    summary_parts = [
        f"Предзаказ по листу «{title}».",
        f"Заявок: {len(preorder_entries)}.",
    ]
    if preview_models:
        summary_parts.append(f"Модели: {preview_models}.")
    if status_summary:
        summary_parts.append(f"Статусы: {status_summary}.")
    summary = " ".join(summary_parts)

    details_parts = [
        f"<p><strong>Лист таблицы:</strong> {clean_inline_html(title)}</p>",
        f"<p><strong>Всего заявок:</strong> {len(preorder_entries)}</p>",
    ]
    if preview_models:
        details_parts.append(
            f"<p><strong>Основные модели:</strong> {clean_inline_html(preview_models)}</p>"
        )
    if status_summary:
        details_parts.append(f"<p><strong>Статусы:</strong> {clean_inline_html(status_summary)}</p>")
    if acquisition_summary:
        details_parts.append(
            f"<p><strong>Способы приобретения:</strong> {clean_inline_html(acquisition_summary)}</p>"
        )

    return {
        "source": source,
        "source_row_key": bound_source_row_key(
            f"{worksheet.title}-{sort_order}",
            fallback=f"worksheet-{sort_order}",
        ),
        "title": title,
        "promotion_kind": Promotion.KIND_PREORDER,
        "badge": "Предзаказ",
        "summary": summary[:220],
        "details": "".join(details_parts),
        "brand": brand,
        "category": "Предзаказы",
        "promo_code": "",
        "cta_label": "",
        "cta_url": "",
        "start_date": min(parsed_dates) if parsed_dates else None,
        "end_date": None,
        "sort_order": sort_order,
        "is_featured": True,
        "is_published": source.auto_publish_imported,
        "raw_data": {
            "Лист таблицы": title,
            "Всего заявок": str(len(preorder_entries)),
            "Модели": ", ".join(unique_models[:6]),
            "Статусы": status_summary,
            "Способы приобретения": acquisition_summary,
            "preorder_entries": preorder_entries,
        },
        "imported_at": timezone.now(),
    }


SIGNATURE_FIELDS = (
    "title",
    "promotion_kind",
    "badge",
    "summary",
    "details",
    "brand",
    "category",
    "promo_price",
    "benefit_value",
    "promo_code",
    "cta_label",
    "cta_url",
    "start_date",
    "end_date",
    "is_featured",
    "is_published",
)


def mapped_promotion_signature(mapped_data):
    signature = []
    for field_name in SIGNATURE_FIELDS:
        value = mapped_data.get(field_name)
        if isinstance(value, (date, datetime)):
            value = value.isoformat()
        signature.append(normalize_header(format_cell_text(value)))
    return tuple(signature)


def deduplicate_mapped_rows(mapped_rows, result):
    unique_rows = []
    rows_by_key = {}
    duplicate_pairs = []
    conflicts = []

    for row_number, mapped_data in mapped_rows:
        row_key = mapped_data["source_row_key"]
        existing = rows_by_key.get(row_key)
        if not existing:
            rows_by_key[row_key] = (row_number, mapped_data)
            unique_rows.append((row_number, mapped_data))
            continue

        existing_row_number, existing_data = existing
        if mapped_promotion_signature(existing_data) == mapped_promotion_signature(mapped_data):
            duplicate_pairs.append((existing_row_number, row_number))
            result.duplicates += 1
            result.skipped += 1
            continue

        conflicts.append(
            {
                "key": row_key,
                "title": mapped_data.get("title", ""),
                "rows": (existing_row_number, row_number),
            }
        )

    if conflicts:
        previews = "; ".join(
            f"строки {item['rows'][0]} и {item['rows'][1]} — {item['title']}"
            for item in conflicts[:5]
        )
        suffix = "" if len(conflicts) <= 5 else f"; и ещё {len(conflicts) - 5}"
        raise ImportValidationError(
            "Обнаружены разные акции с одинаковым ключом импорта: "
            f"{previews}{suffix}. Добавьте им разные значения в колонке «ID» или «Код акции»."
        )

    if duplicate_pairs:
        pair_preview = ", ".join(
            f"{first}/{second}"
            for first, second in duplicate_pairs[:8]
        )
        result.warnings.append(
            f"Повторяющиеся строки пропущены: {result.duplicates} (пары строк: {pair_preview})."
        )

    return unique_rows


def upsert_mapped_promotion(source, mapped_data, result, seen_keys):
    seen_keys.append(mapped_data["source_row_key"])

    existing = (
        Promotion.objects.filter(source=source, source_row_key=mapped_data["source_row_key"])
        .order_by("pk")
        .first()
    )

    if existing and not existing.sync_with_source:
        result.skipped += 1
        return

    if existing:
        for field_name, field_value in mapped_data.items():
            if (
                field_name in {"promotion_kind", "start_date", "end_date"}
                and not field_value
                and getattr(existing, field_name)
            ):
                continue
            setattr(existing, field_name, field_value)
        existing.save()
        result.updated += 1
        return

    Promotion.objects.create(**mapped_data)
    result.created += 1


def preview_mapped_promotions(source, mapped_rows, result):
    seen_keys = []
    existing_by_key = {}
    for promotion in Promotion.objects.filter(source=source).order_by("pk"):
        existing_by_key.setdefault(promotion.source_row_key, promotion)

    for _row_number, mapped_data in mapped_rows:
        row_key = mapped_data["source_row_key"]
        seen_keys.append(row_key)
        existing = existing_by_key.get(row_key)
        if existing and not existing.sync_with_source:
            result.skipped += 1
        elif existing:
            result.updated += 1
        else:
            result.created += 1

    if source.archive_missing_on_import:
        result.unpublished = (
            Promotion.objects.filter(source=source, sync_with_source=True, is_published=True)
            .exclude(source_row_key__in=seen_keys)
            .count()
        )

    return result


def validate_import_volume(source, mapped_rows):
    """Останавливает импорт до записи, если выгрузка выглядит неполной."""
    imported_keys = {
        mapped_data["source_row_key"]
        for _row_number, mapped_data in mapped_rows
    }
    imported_count = len(imported_keys)

    if imported_count < source.minimum_expected_rows:
        raise ImportValidationError(
            f"Распознано акций: {imported_count}, а для источника установлен минимум "
            f"{source.minimum_expected_rows}. Импорт остановлен: проверьте доступность, "
            "лист и строку заголовков."
        )

    if not source.pk or not source.archive_missing_on_import:
        return

    existing_keys = set(
        Promotion.objects.filter(source=source, sync_with_source=True)
        .exclude(source_row_key="")
        .values_list("source_row_key", flat=True)
    )
    if not existing_keys:
        return

    missing_count = len(existing_keys - imported_keys)
    missing_percent = (
        missing_count * 100 + len(existing_keys) - 1
    ) // len(existing_keys)
    if missing_percent > source.max_missing_percent:
        raise ImportValidationError(
            f"Из {len(existing_keys)} ранее синхронизированных акций пропало "
            f"{missing_count} ({missing_percent}%), допустимый предел — "
            f"{source.max_missing_percent}%. Импорт остановлен, чтобы не снять акции "
            "с публикации из-за неполной выгрузки. Проверьте таблицу или временно "
            "увеличьте предел в настройках источника."
        )


def persist_mapped_promotions(source, mapped_rows, result):
    seen_keys = []
    with transaction.atomic():
        if source.pk:
            # PostgreSQL блокирует одновременную запись двух импортов одного источника.
            source = PromotionSource.objects.select_for_update().get(pk=source.pk)
        for _row_number, mapped_data in mapped_rows:
            upsert_mapped_promotion(source, mapped_data, result, seen_keys)
        finalize_source_import(source, seen_keys, result)
    return result


def finalize_source_import(source, seen_keys, result):
    if source.archive_missing_on_import:
        result.unpublished = (
            Promotion.objects.filter(source=source, sync_with_source=True, is_published=True)
            .exclude(source_row_key__in=seen_keys)
            .update(is_published=False)
        )

    source.last_imported_at = timezone.now()
    source.last_import_error = ""
    source.save(update_fields=["last_imported_at", "last_import_error", "updated_at"])


def prepare_preorders_from_last_worksheets(source):
    payload, _ = load_payload(source.xlsx_url)

    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"Не удалось прочитать Excel-книгу: {exc}") from exc

    result = ImportResult()
    mapped_rows = []

    try:
        worksheet_count = max(source.worksheets_to_import, 1)
        worksheets = workbook.worksheets[-worksheet_count:]

        for index, worksheet in enumerate(worksheets, start=1):
            mapped_data = map_worksheet_to_preorder_promotion(
                source,
                worksheet,
                index * 10,
            )
            if not mapped_data:
                result.skipped += 1
                continue

            mapped_rows.append((index, mapped_data))

        return deduplicate_mapped_rows(mapped_rows, result), result
    finally:
        workbook.close()


def prepare_single_sheet_promotions(source):
    result = ImportResult()
    mapped_rows = []
    generated_key_rows = 0
    overridden_rows = 0
    rows = fetch_source_rows(source)

    for row_number, raw_row in rows:
        mapped_data = map_row_to_promotion(source, row_number, raw_row)
        if not mapped_data:
            result.skipped += 1
            continue
        validate_mapped_promotion_dates(row_number, raw_row, mapped_data)
        normalized_row = {
            normalize_header(key): value.strip()
            for key, value in raw_row.items()
            if key
        }
        if not extract_value(normalized_row, "row_key"):
            generated_key_rows += 1
        if mapped_data.get("raw_data", {}).get("Цвет (правило импорта)"):
            overridden_rows += 1
        mapped_rows.append((row_number, mapped_data))

    if generated_key_rows:
        result.warnings.append(
            f"У {generated_key_rows} строк нет явного ID: ключ рассчитан из товара, бренда и дат. "
            "Чтобы изменение периода обновляло существующую акцию, добавьте колонку «ID»."
        )
    if overridden_rows:
        result.warnings.append(
            f"Для {overridden_rows} строк применены локальные правила дополнения данных, "
            "отсутствующих в Google-таблице."
        )

    return deduplicate_mapped_rows(mapped_rows, result), result


def record_source_import_error(source, error):
    if not source.pk:
        return
    source.last_import_error = str(error)
    source.save(update_fields=["last_import_error", "updated_at"])


def start_import_run(source, *, dry_run):
    if not source.pk:
        return None
    return PromotionImportRun.objects.create(
        source=source,
        is_dry_run=dry_run,
    )


def finish_import_run(import_run, *, result=None, error=None):
    if not import_run:
        return

    import_run.finished_at = timezone.now()
    update_fields = ["status", "finished_at"]
    if error is not None:
        import_run.status = PromotionImportRun.STATUS_ERROR
        import_run.error = str(error)
        update_fields.append("error")
    else:
        import_run.status = PromotionImportRun.STATUS_SUCCESS
        import_run.created_count = result.created
        import_run.updated_count = result.updated
        import_run.skipped_count = result.skipped
        import_run.unpublished_count = result.unpublished
        import_run.duplicate_count = result.duplicates
        import_run.warnings = list(result.warnings)
        update_fields.extend(
            [
                "created_count",
                "updated_count",
                "skipped_count",
                "unpublished_count",
                "duplicate_count",
                "warnings",
            ]
        )
    import_run.save(update_fields=update_fields)


def import_preorders_from_last_worksheets(source, *, dry_run=False):
    mapped_rows, result = prepare_preorders_from_last_worksheets(source)
    validate_import_volume(source, mapped_rows)
    if dry_run:
        return preview_mapped_promotions(source, mapped_rows, result)
    return persist_mapped_promotions(source, mapped_rows, result)


def import_promotions_from_source(source, *, dry_run=False):
    import_run = start_import_run(source, dry_run=dry_run)
    try:
        if source.import_mode == PromotionSource.IMPORT_MODE_LAST_WORKSHEETS:
            result = import_preorders_from_last_worksheets(source, dry_run=dry_run)
        else:
            mapped_rows, result = prepare_single_sheet_promotions(source)
            validate_import_volume(source, mapped_rows)
            if dry_run:
                result = preview_mapped_promotions(source, mapped_rows, result)
            else:
                result = persist_mapped_promotions(source, mapped_rows, result)

    except Exception as exc:
        if not dry_run:
            record_source_import_error(source, exc)
        finish_import_run(import_run, error=exc)
        raise

    finish_import_run(import_run, result=result)
    return result


def preview_promotions_from_source(source):
    return import_promotions_from_source(source, dry_run=True)
